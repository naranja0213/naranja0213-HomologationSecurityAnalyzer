"""Excel report generation."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ALL_FILES_HEADERS, HASH_REPORT_HEADERS


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
APPROVED_FILL = PatternFill("solid", fgColor="C6EFCE")
MANUAL_FILL = PatternFill("solid", fgColor="FFF2CC")
HIGH_RISK_FILL = PatternFill("solid", fgColor="FFC7CE")


def create_excel_report(
    output_path: Path,
    summary: dict[str, Any],
    hash_rows: list[dict[str, Any]],
    all_file_rows: list[dict[str, Any]],
) -> None:
    """Create the workbook required by the homologation workflow."""

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, summary)

    hash_sheet = workbook.create_sheet("Hash Report")
    _write_table(hash_sheet, HASH_REPORT_HEADERS, hash_rows)

    detected_sheet = workbook.create_sheet("Detected Files")
    detected_rows = [row for row in hash_rows if _as_int(row.get("VT_Malicious")) > 0]
    _write_table(detected_sheet, HASH_REPORT_HEADERS, detected_rows)

    unsigned_sheet = workbook.create_sheet("Unsigned Files")
    unsigned_rows = [row for row in hash_rows if row.get("SignatureStatus") == "NotSigned"]
    _write_table(unsigned_sheet, HASH_REPORT_HEADERS, unsigned_rows)

    all_files_sheet = workbook.create_sheet("All Files")
    _write_table(all_files_sheet, ALL_FILES_HEADERS, all_file_rows)

    for sheet in workbook.worksheets:
        _auto_size_columns(sheet)
        _style_decisions(sheet)

    workbook.save(output_path)


def _write_summary(sheet, summary: dict[str, Any]) -> None:
    sheet.append(["Field", "Value"])
    for key, value in summary.items():
        sheet.append([key, value])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def _write_table(sheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def _auto_size_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in column_cells[1:]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))

        width = max(10, min(max_length + 2, 80))
        if header in {"SHA256"}:
            width = 68
        elif header in {"RelativePath", "VT_DetectedVendors", "GLPI_Comment"}:
            width = 60
        sheet.column_dimensions[column_letter].width = width


def _style_decisions(sheet) -> None:
    headers = [cell.value for cell in sheet[1]]
    decision_index = _header_index(headers, "Decision")
    risk_index = _header_index(headers, "RiskLevel")

    for row in sheet.iter_rows(min_row=2):
        for index in [decision_index, risk_index]:
            if index is None:
                continue
            cell = row[index]
            text = str(cell.value or "").lower()
            if "reject" in text or "high" in text:
                cell.fill = HIGH_RISK_FILL
            elif "manual" in text or "medium" in text or "unknown" in text:
                cell.fill = MANUAL_FILL
            elif "approved" in text or "low" in text:
                cell.fill = APPROVED_FILL


def _header_index(headers: list[Any], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _as_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0

